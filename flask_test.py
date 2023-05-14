import openpyxl
from flask import Flask, render_template

# 엑셀 파일 열기
workbook = openpyxl.load_workbook('C:/Users/황지혁/Desktop/flask_test/flask_excel.xlsx')
# 첫번째 시트 선택
sheet = workbook.active
# 첫번째, 두번째 열의 모든 값을 가져와 배열에 저장
column1 = []
column2 = []
for row in sheet.iter_rows(min_row=1, min_col=1, max_col=2, values_only=True):
    column1.append(row[0])
    column2.append(row[1])

# Flask 앱 생성
app = Flask(__name__)

@app.route('/')
def home():
    return render_template('home.html')

# 첫 번째 메뉴
@app.route('/menu1')
def menu1():
    return render_template('menu1.html', data=column1)
# 두 번째 메뉴
@app.route('/menu2')
def menu2():
    return render_template('menu2.html', data=column2)

# 업데이트된 엑셀 받아오기 - menu1
@app.route('/menu1/update', methods=['POST'])
def menu1_update():
    workbook = openpyxl.load_workbook('C:/Users/황지혁/Desktop/flask_test/flask_excel.xlsx')
    sheet = workbook.active
    column1 = []
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        column1.append(row[0])
    return render_template('menu1.html', data=column1)

# 업데이트된 엑셀 받아오기 - menu2
@app.route('/menu2/update', methods=['POST'])
def menu2_update():
    workbook = openpyxl.load_workbook('C:/Users/황지혁/Desktop/flask_test/flask_excel.xlsx')
    sheet = workbook.active
    column2 = []
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=2, values_only=True):
        column2.append(row[1])
    return render_template('menu2.html', data=column2)


if __name__ == '__main__':
    app.run()


